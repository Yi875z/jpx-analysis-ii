#!/usr/bin/env python3
"""
build_report.py
JPX分析結果からMarkdown／Excel／PDFレポートを生成する。

使用例:
  python scripts/build_report.py --analysis /tmp/jpx_analysis.json --format all
  python scripts/build_report.py --analysis /tmp/jpx_analysis.json --format md
  python scripts/build_report.py --analysis /tmp/jpx_analysis.json --format xlsx
  python scripts/build_report.py --analysis /tmp/jpx_analysis.json --format pdf
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

OUTPUT_DIR = Path("/mnt/user-data/outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

INVESTOR_LABELS = {
    "foreign": "海外投資家",
    "individual": "個人投資家",
    "trust": "投信・信託銀行",
    "corporate": "事業法人",
    "dealer": "自己（証券会社）",
}


def load_analysis(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def fmt_net(val):
    if val is None:
        return "-"
    mark = "▲" if val > 0 else "▼"
    return f"{mark}{abs(val):,.0f}"


def fmt_diff(val):
    if val is None:
        return "-"
    return f"{'+' if val >= 0 else ''}{val:,.0f}"


def build_markdown(analysis: dict, mode: str = "weekly") -> str:
    """Markdownレポート本文を生成"""
    spot = analysis.get("spot", {})
    futures = analysis.get("futures", {})

    date_str = spot.get("latest_date", futures.get("latest_date", datetime.now().strftime("%Y-%m-%d")))
    title_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y年%m月%d日") if date_str != "不明" else "最新"

    lines = []
    lines.append(f"# JPX投資家別売買動向 {title_date}週")
    lines.append(f"> データソース: JPX投資部門別売買状況 | 集計基準: {date_str}")
    lines.append("")

    # エグゼクティブサマリー（自動生成）
    lines.append("## 📋 エグゼクティブサマリー")
    lines.append("")
    if spot.get("investors"):
        foreign = next((i for i in spot["investors"] if i["key"] == "foreign"), None)
        individual = next((i for i in spot["investors"] if i["key"] == "individual"), None)
        if foreign and foreign["net_latest"] is not None:
            direction = "買い越し" if foreign["net_latest"] > 0 else "売り越し"
            lines.append(f"- 海外投資家は現物で **{fmt_net(foreign['net_latest'])}億円の{direction}**（前週比{fmt_diff(foreign['week_diff'])}億円）。")
        if individual and individual["net_latest"] is not None:
            direction = "買い越し" if individual["net_latest"] > 0 else "売り越し"
            lines.append(f"- 個人投資家は現物で{fmt_net(individual['net_latest'])}億円の{direction}（逆張り動向{'あり' if (foreign and individual['net_latest'] * foreign['net_latest'] < 0) else 'なし'}）。")
    lines.append("")

    # 現物テーブル
    if spot.get("investors"):
        lines.append("## 📊 現物（東証プライム）：投資家別買い越し・売り越し")
        lines.append("")
        lines.append("| 投資家区分 | 差引(億円) | 前週比 | 判定 |")
        lines.append("|-----------|----------|-------|------|")
        for inv in spot["investors"]:
            net = inv["net_latest"]
            judge = ("買い越し" if net > 0 else "売り越し") if net is not None else "-"
            lines.append(f"| {inv['label']} | {fmt_net(net)} | {fmt_diff(inv['week_diff'])} | {judge} |")
        lines.append("")

    # 先物テーブル
    if futures.get("investors"):
        lines.append("## 📈 先物（日経225・TOPIX）：投資家別動向")
        lines.append("")
        lines.append("| 投資家区分 | ネット(枚) | 前週比 | 判定 |")
        lines.append("|-----------|----------|-------|------|")
        for inv in futures["investors"]:
            net = inv["net_latest"]
            judge = ("買建超" if net > 0 else "売建超") if net is not None else "-"
            lines.append(f"| {inv['label']} | {fmt_net(net)} | {fmt_diff(inv['week_diff'])} | {judge} |")
        lines.append("")

    # 注目セグメント
    lines.append("## 🔍 注目セグメント動向")
    lines.append("")
    if spot.get("investors"):
        for key in ["foreign", "individual", "trust"]:
            inv = next((i for i in spot["investors"] if i["key"] == key), None)
            if inv and inv["net_latest"] is not None:
                label = INVESTOR_LABELS[key]
                direction = "買い越し" if inv["net_latest"] > 0 else "売り越し"
                lines.append(f"### {'🔵' if key == 'foreign' else '🟡' if key == 'individual' else '🟢'} {label}")
                lines.append(f"現物で{fmt_net(inv['net_latest'])}億円の{direction}（前週比{fmt_diff(inv['week_diff'])}億円）。")
                lines.append("")

    # 前週比・前月比テーブル
    lines.append("## 📅 先週比変化")
    lines.append("")
    lines.append("| 投資家区分 | 先週比(億円) |")
    lines.append("|-----------|-----------|")
    for inv in spot.get("investors", []):
        lines.append(f"| {inv['label']} | {fmt_diff(inv['week_diff'])} |")
    lines.append("")

    lines.append("---")
    lines.append(f"*生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')} | スキル: jpx-investor-data*")

    return "\n".join(lines)


def save_markdown(content: str, date_str: str) -> Path:
    filename = f"jpx_investor_{date_str.replace('-', '')}.md"
    path = OUTPUT_DIR / filename
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"[Markdown] 保存: {path}")
    return path


def build_excel(analysis: dict, date_str: str) -> Path:
    """Excelファイルを生成（openpyxl使用）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    title_font = Font(bold=True, size=11, name="Arial")

    def add_investor_sheet(wb, data: dict, sheet_name: str):
        ws = wb.create_sheet(sheet_name)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        # ヘッダー
        headers = ["投資家区分", "差引(億円)", "前週比", "判定"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # データ行
        for row, inv in enumerate(data.get("investors", []), 2):
            net = inv["net_latest"]
            diff = inv["week_diff"]
            judge = ("買い越し" if net > 0 else "売り越し") if net is not None else "-"

            ws.cell(row=row, column=1, value=inv["label"]).font = Font(name="Arial")
            ws.cell(row=row, column=2, value=round(net, 0) if net else None)
            ws.cell(row=row, column=3, value=round(diff, 0) if diff else None)
            ws.cell(row=row, column=4, value=judge)

            # 買い越し=青、売り越し=赤
            if net is not None:
                color = "1F4E79" if net > 0 else "C00000"
                ws.cell(row=row, column=2).font = Font(color=color, name="Arial", bold=True)

    spot = analysis.get("spot", {})
    futures = analysis.get("futures", {})

    if spot.get("investors"):
        add_investor_sheet(wb, spot, "現物")
    if futures.get("investors"):
        add_investor_sheet(wb, futures, "先物")

    # 合算シート（現物＋先物）
    if spot.get("investors") and futures.get("investors"):
        ws_sum = wb.create_sheet("合算")
        ws_sum.cell(row=1, column=1, value="合算（現物＋先物換算）").font = title_font
        # 簡易合算
        ws_sum.append(["投資家区分", "現物ネット", "先物ネット換算", "合算"])
        for s_inv, f_inv in zip(spot["investors"], futures["investors"]):
            s_net = s_inv["net_latest"] or 0
            f_net = f_inv["net_latest"] or 0
            ws_sum.append([s_inv["label"], s_net, f_net, s_net + f_net])

    filename = f"jpx_investor_{date_str.replace('-', '')}.xlsx"
    path = OUTPUT_DIR / filename
    wb.save(path)
    print(f"[Excel] 保存: {path}")
    return path


def build_pdf(md_content: str, date_str: str) -> Path:
    """MarkdownからPDFを生成"""
    # まずMarkdownを一時保存
    md_path = Path("/tmp") / f"jpx_tmp_{date_str}.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    filename = f"jpx_investor_{date_str.replace('-', '')}.pdf"
    pdf_path = OUTPUT_DIR / filename

    # pandocが使えれば最良
    try:
        result = subprocess.run(
            ["pandoc", str(md_path), "-o", str(pdf_path),
             "--pdf-engine=weasyprint", "-V", "lang=ja"],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0:
            print(f"[PDF] 保存: {pdf_path}")
            return pdf_path
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    # reportlabフォールバック
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "reportlab", "--break-system-packages", "-q"])
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        for line in md_content.split("\n"):
            line = line.strip()
            if line.startswith("# "):
                story.append(Paragraph(line[2:], styles["Title"]))
            elif line.startswith("## "):
                story.append(Paragraph(line[3:], styles["Heading2"]))
            elif line:
                # Markdownの記号を簡易除去
                clean = line.replace("|", " ").replace("▲", "+").replace("▼", "-")
                story.append(Paragraph(clean, styles["Normal"]))
            story.append(Spacer(1, 4))
        doc.build(story)
        print(f"[PDF] 保存（reportlab）: {pdf_path}")
        return pdf_path
    except Exception as e:
        print(f"[警告] PDF生成失敗: {e}。Markdownファイルを代替として使用してください。")
        return None


def main():
    parser = argparse.ArgumentParser(description="JPXレポート生成")
    parser.add_argument("--analysis", required=True, help="analyze_jpx.py の出力JSONパス")
    parser.add_argument("--format", choices=["md", "xlsx", "pdf", "all"], default="all")
    parser.add_argument("--mode", choices=["weekly", "monthly"], default="weekly")
    args = parser.parse_args()

    analysis = load_analysis(args.analysis)

    # 基準日を取得
    spot = analysis.get("spot", {})
    date_str = spot.get("latest_date", datetime.now().strftime("%Y-%m-%d"))

    outputs = []

    if args.format in ("md", "all"):
        md_content = build_markdown(analysis, args.mode)
        md_path = save_markdown(md_content, date_str)
        outputs.append(md_path)

    if args.format in ("xlsx", "all"):
        xlsx_path = build_excel(analysis, date_str)
        if xlsx_path:
            outputs.append(xlsx_path)

    if args.format in ("pdf", "all"):
        if "md_content" not in dir():
            md_content = build_markdown(analysis, args.mode)
        pdf_path = build_pdf(md_content, date_str)
        if pdf_path:
            outputs.append(pdf_path)

    print(f"\n=== 出力完了 ===")
    for p in outputs:
        print(f"  {p}")

    return outputs


if __name__ == "__main__":
    main()
