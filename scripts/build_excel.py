"""
scripts/build_excel.py
Supabase蓄積データからExcel累積ファイルを生成する
"""

import io
import logging
from datetime import date, datetime
from pathlib import Path
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")   # GUI不要モード
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
from matplotlib import rcParams
rcParams["font.family"] = "Meiryo"   # 日本語フォント（Windows）

logger = logging.getLogger(__name__)

INVESTORS = ["foreign", "individual", "trust_bank", "inv_trust", "corporate", "dealer"]
INVESTOR_JP = {
    "foreign":    "海外投資家",
    "individual": "個人投資家",
    "trust_bank": "信託銀行",
    "inv_trust":  "投資信託",
    "corporate":  "事業法人",
    "dealer":     "自己（証券会社）",
}

# カラーコード（xlsxスキル準拠）
COLOR_HEADER = "1F4E79"   # 紺
COLOR_BUY    = "1F4E79"   # 買い越し=青
COLOR_SELL   = "C00000"   # 売り越し=赤
COLOR_SUBHD  = "D6E4F0"   # サブヘッダー薄青
COLOR_TWIN   = "FFD700"   # 両輪買い=金

def _header_style(cell, bg=COLOR_HEADER):
    cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _set_border(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _calc_rolling_zscores(history: list[dict], window: int = 52) -> dict:
    """
    spot_history（全投資家区分混在可）から investor_type ごとに
    week_date → zscore のマップを生成する。
    各週のZスコアはその週までの過去window週データで計算（ローリング）。
    """
    from collections import defaultdict
    by_inv = defaultdict(list)
    for r in history:
        if r.get("net_amount") is not None:
            by_inv[r["investor_type"]].append((r["week_date"], r["net_amount"]))

    result = {}   # (week_date, investor_type) → zscore
    for inv, rows in by_inv.items():
        sorted_rows = sorted(rows, key=lambda x: x[0])  # 古い順
        nets = [v for _, v in sorted_rows]
        for i, (wd, _) in enumerate(sorted_rows):
            start = max(0, i - window + 1)
            window_nets = nets[start : i + 1]
            if len(window_nets) < 4:
                result[(wd, inv)] = None
                continue
            arr = np.array(window_nets, dtype=float)
            mean, std = arr.mean(), arr.std()
            if std == 0:
                result[(wd, inv)] = 0.0
            else:
                result[(wd, inv)] = round(float((nets[i] - mean) / std), 3)
    return result


def _zscore_fill(z):
    """Zスコアの大きさに応じた背景色を返す"""
    if z is None:
        return None
    if z >= 2.0:  return "1F4E79"   # 濃青（強い買い超過）
    if z >= 1.0:  return "9DC3E6"   # 淡青
    if z <= -2.0: return "C00000"   # 濃赤（強い売り超過）
    if z <= -1.0: return "F4CCCC"   # 淡赤
    return "F2F2F2"                  # 中立グレー


def _write_spot_sheet(ws, spot_history: list[dict], futures_history: list[dict],
                      zscore_history: list[dict] = None):
    """現物週次シートを書き込む（ネット額＋Zスコアカラム付き）
    zscore_history: Zスコア計算用の全履歴（年をまたいだ52週窓のため spot_history より広い範囲）
    """
    ws.title = "現物_週次"

    # ヘッダー：ネット列（A〜H）＋区切り＋Zスコア列（J〜O）
    net_headers = ["基準週", "海外投資家", "個人投資家", "信託銀行",
                   "投資信託", "事業法人", "自己（証券）", "合計ネット"]
    for c, h in enumerate(net_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(c)].width = 13

    # 区切り列 I
    ws.cell(row=1, column=9, value="").fill = PatternFill("solid", fgColor="FFFFFF")
    ws.column_dimensions["I"].width = 3

    # Zスコアヘッダー（J〜O：海外・個人・信託銀行・投信・事業法人・自己）
    z_headers = ["Z:海外", "Z:個人", "Z:信託銀", "Z:投信", "Z:事業法人", "Z:自己"]
    for c, h in enumerate(z_headers, 10):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell, bg="4A4A4A")
        ws.column_dimensions[get_column_letter(c)].width = 9

    # Zスコア: 全履歴（年をまたぐ）からローリング計算（weekly_statsに依存しない）
    base = zscore_history if zscore_history else spot_history
    stats_map = _calc_rolling_zscores(base, window=52)

    # データをグループ化
    by_week = {}
    for r in spot_history:
        wd = r["week_date"]
        by_week.setdefault(wd, {})[r["investor_type"]] = r.get("net_amount", 0) or 0

    for row_i, (wd, inv_data) in enumerate(sorted(by_week.items()), 2):
        ws.cell(row=row_i, column=1, value=wd).font = Font(name="Arial", size=9)
        total = 0
        for col_i, inv in enumerate(INVESTORS, 2):
            val = inv_data.get(inv, 0) or 0
            total += val
            cell = ws.cell(row=row_i, column=col_i, value=round(val, 2))
            cell.font = Font(color=COLOR_BUY if val >= 0 else COLOR_SELL,
                             name="Arial", size=9, bold=(abs(val) > 1000))
            cell.number_format = '#,##0.0;[Red]-#,##0.0'
            _set_border(cell)
        ws.cell(row=row_i, column=8, value=round(total, 2)).number_format = '#,##0.0'

        # Zスコア列（J〜O）
        for col_i, inv in enumerate(INVESTORS, 10):
            z = stats_map.get((wd, inv))
            cell = ws.cell(row=row_i, column=col_i,
                           value=round(z, 2) if z is not None else "")
            cell.font = Font(name="Arial", size=9,
                             color="FFFFFF" if z is not None and abs(z) >= 2.0 else "000000")
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = "0.00"
            bg = _zscore_fill(z)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            _set_border(cell)

    logger.info(f"[Excel] 現物シート: {len(by_week)}週分")


def _make_chart_image(spot_history: list[dict], investor_type: str,
                      title: str, width_inch: float = 14, height_inch: float = 6) -> io.BytesIO:
    """matplotlibで棒グラフ（買い越し=青/売り越し=赤）＋4週MA線を生成してBytesIOで返す"""
    rows = sorted(
        [r for r in spot_history if r["investor_type"] == investor_type],
        key=lambda x: x["week_date"]
    )
    if not rows:
        return None

    dates = [datetime.strptime(r["week_date"], "%Y-%m-%d") for r in rows]
    vals  = [r.get("net_amount", 0) or 0 for r in rows]

    # 4週移動平均
    ma4 = []
    for i in range(len(vals)):
        w = vals[max(0, i - 3): i + 1]
        ma4.append(sum(w) / len(w))

    colors = ["#1F4E79" if v >= 0 else "#C00000" for v in vals]

    fig, ax = plt.subplots(figsize=(width_inch, height_inch))
    fig.patch.set_facecolor("#FAFAFA")
    ax.set_facecolor("#FAFAFA")

    # 棒グラフ
    bars = ax.bar(dates, vals, color=colors, width=5, edgecolor="none", alpha=0.85)

    # 4週MA線
    ax.plot(dates, ma4, color="#FF8C00", linewidth=2.0, label="4週移動平均", zorder=5)

    # ゼロライン
    ax.axhline(0, color="#333333", linewidth=0.8, linestyle="-")

    # 軸フォーマット
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%y/%m/%d"))
    n = len(dates)
    skip = max(1, n // 15)
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=skip))
    plt.xticks(rotation=45, ha="right", fontsize=8)

    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{int(x):,}" if abs(x) >= 1 else f"{x:.1f}"
    ))
    ax.set_ylabel("億円", fontsize=10)
    ax.yaxis.set_tick_params(labelsize=9)

    # グリッド（横のみ）
    ax.yaxis.grid(True, linestyle="--", linewidth=0.5, color="#CCCCCC", alpha=0.8)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    # タイトル・凡例
    ax.set_title(title, fontsize=13, fontweight="bold", pad=12)
    buy_patch  = plt.Rectangle((0, 0), 1, 1, fc="#1F4E79", label="買い越し")
    sell_patch = plt.Rectangle((0, 0), 1, 1, fc="#C00000", label="売り越し")
    ma_line    = plt.Line2D([0], [0], color="#FF8C00", linewidth=2, label="4週移動平均")
    ax.legend(handles=[buy_patch, sell_patch, ma_line],
              loc="upper left", fontsize=9, framealpha=0.9)

    fig.tight_layout(pad=1.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_comparison_image(spot_history: list[dict],
                           width_inch: float = 14, height_inch: float = 7) -> io.BytesIO:
    """全投資家ネット比較ラインチャートをBytesIOで返す"""
    inv_order  = ["foreign", "individual", "trust_bank", "inv_trust", "corporate", "dealer"]
    inv_labels = {k: INVESTOR_JP[k] for k in inv_order}
    inv_colors = {
        "foreign":    "#1F4E79",
        "individual": "#C00000",
        "trust_bank": "#70AD47",
        "inv_trust":  "#44B3A1",
        "corporate":  "#FF8C00",
        "dealer":     "#7030A0",
    }
    inv_widths = {"foreign": 2.5}  # 海外は太く

    by_week = {}
    for r in spot_history:
        wd = r["week_date"]
        by_week.setdefault(wd, {})[r["investor_type"]] = r.get("net_amount", 0) or 0

    dates = sorted(by_week.keys())
    dt_dates = [datetime.strptime(d, "%Y-%m-%d") for d in dates]
    n = len(dates)

    fig, ax = plt.subplots(figsize=(width_inch, height_inch))
    fig.patch.set_facecolor("#FAFAFA")
    ax.set_facecolor("#FAFAFA")

    for inv in inv_order:
        vals = [by_week[d].get(inv, 0) for d in dates]
        ax.plot(dt_dates, vals,
                color=inv_colors[inv],
                linewidth=inv_widths.get(inv, 1.5),
                label=inv_labels[inv],
                alpha=0.9)

    ax.axhline(0, color="#333333", linewidth=0.8, linestyle="-")

    skip = max(1, n // 15)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%y/%m/%d"))
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=skip))
    plt.xticks(rotation=45, ha="right", fontsize=8)

    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{int(x):,}"
    ))
    ax.set_ylabel("億円", fontsize=10)
    ax.yaxis.grid(True, linestyle="--", linewidth=0.5, color="#CCCCCC", alpha=0.8)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

    ax.set_title("全投資家 現物週次ネット比較（億円）", fontsize=13, fontweight="bold", pad=12)
    ax.legend(loc="upper left", fontsize=9, framealpha=0.9, ncol=2)

    fig.tight_layout(pad=1.5)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def _write_chart_sheet(wb, spot_history: list[dict]):
    """外国人ネット推移チャートシート（matplotlib画像を埋め込み）"""
    ws = wb.create_sheet("チャート_外国人")
    ws.sheet_view.showGridLines = False

    buf = _make_chart_image(
        spot_history, "foreign",
        "海外投資家 現物週次ネット（億円）\n■青=買い越し  ■赤=売り越し  ─オレンジ=4週移動平均",
    )
    if buf:
        img = XLImage(buf)
        img.anchor = "A1"
        ws.add_image(img)

    # 全投資家比較チャートシートも追加
    ws2 = wb.create_sheet("チャート_全投資家")
    ws2.sheet_view.showGridLines = False
    buf2 = _make_comparison_image(spot_history)
    if buf2:
        img2 = XLImage(buf2)
        img2.anchor = "A1"
        ws2.add_image(img2)


def _write_monthly_sheet(ws, monthly_data: list[dict]):
    """月次サマリーシート"""
    ws.title = "月次サマリー"
    ws.column_dimensions["A"].width = 12
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 15

    headers = ["年月", "海外投資家", "個人投資家", "信託銀行", "投資信託", "事業法人", "自己（証券）", "集計週数"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell)

    by_month = {}
    for r in monthly_data:
        ym = r["year_month"]
        if ym not in by_month:
            by_month[ym] = {}
        by_month[ym][r["investor_type"]] = r

    for row_i, (ym, inv_data) in enumerate(sorted(by_month.items(), reverse=True), 2):
        ws.cell(row=row_i, column=1, value=ym).font = Font(name="Arial", size=9)
        for col_i, inv in enumerate(INVESTORS, 2):
            d = inv_data.get(inv, {})
            val = d.get("combined_net", 0) or 0
            cell = ws.cell(row=row_i, column=col_i, value=round(val, 2))
            cell.font = Font(
                color=COLOR_BUY if val >= 0 else COLOR_SELL,
                name="Arial", size=9
            )
            cell.number_format = '#,##0.0'
        wc = next(iter(inv_data.values()), {}).get("week_count", 0)
        ws.cell(row=row_i, column=8, value=wc)


def _make_zscore_image(spot_history: list[dict],
                       futures_history: list[dict] = None,
                       width_inch: float = 14, height_inch: float = 12) -> io.BytesIO:
    """海外投資家のZスコア推移チャート（52週ローリング）をBytesIOで返す
    上段：現物Zスコア / 下段：先物Zスコア（futures_history が渡された場合）
    """
    foreign_spot = sorted(
        [r for r in spot_history if r.get("investor_type") == "foreign"],
        key=lambda x: x["week_date"]
    )
    if len(foreign_spot) < 4:
        return None

    # 先物データ（nikkei225_large + topix_large 合算）
    has_futures = False
    if futures_history:
        from collections import defaultdict
        fut_by_week = defaultdict(float)
        for r in futures_history:
            if r.get("investor_type") == "foreign":
                fut_by_week[r["week_date"]] += (r.get("net_amount_oku") or 0)
        if fut_by_week:
            has_futures = True
            fut_rows = [{"week_date": d, "net_amount_oku": v, "investor_type": "foreign"}
                        for d, v in sorted(fut_by_week.items())]

    nrows = 2 if has_futures else 1
    actual_height = height_inch if has_futures else height_inch / 2
    fig, axes = plt.subplots(nrows, 1, figsize=(width_inch, actual_height),
                             squeeze=False)
    fig.patch.set_facecolor("#FAFAFA")

    def _draw_zscore_panel(ax, rows, value_key, title):
        """共通の描画処理"""
        ax.set_facecolor("#FAFAFA")
        zmap52 = _calc_rolling_zscores(rows, window=52)
        zmap26 = _calc_rolling_zscores(rows, window=26)
        dates = [datetime.strptime(r["week_date"], "%Y-%m-%d") for r in rows]
        z52 = [zmap52.get((r["week_date"], "foreign")) for r in rows]
        z26 = [zmap26.get((r["week_date"], "foreign")) for r in rows]

        ax.axhspan(1, 3.5,   alpha=0.08, color="#1F4E79", label="_nolegend_")
        ax.axhspan(-3.5, -1, alpha=0.08, color="#C00000", label="_nolegend_")
        for y, ls in [(0, "-"), (1, "--"), (-1, "--"), (2, ":"), (-2, ":")]:
            ax.axhline(y, color="#888888", linewidth=0.7, linestyle=ls)

        z52_clean = [v if v is not None else 0.0 for v in z52]
        colors_z = ["#1F4E79" if v >= 0 else "#C00000" for v in z52_clean]
        ax.bar(dates, z52_clean, color=colors_z, width=5, alpha=0.7, label="Zスコア(52週)")

        z26_clean = [v if v is not None else float("nan") for v in z26]
        ax.plot(dates, z26_clean, color="#FF8C00", linewidth=1.8,
                linestyle="--", label="Zスコア(26週)", zorder=5)

        skip = max(1, len(dates) // 15)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%y/%m/%d"))
        ax.xaxis.set_major_locator(mdates.WeekdayLocator(interval=skip))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=8)

        ax.set_ylabel("Zスコア（標準偏差）", fontsize=10)
        ax.set_ylim(-3.5, 3.5)
        ax.yaxis.set_tick_params(labelsize=9)
        ax.yaxis.grid(False)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.text(0.01, 0.97, "+2以上：統計的に強い買い超過  |  -2以下：統計的に強い売り超過",
                transform=ax.transAxes, fontsize=8, color="#555555", va="top")
        ax.set_title(title, fontsize=13, fontweight="bold", pad=12)
        ax.legend(loc="upper left", fontsize=9, framealpha=0.9)

    _draw_zscore_panel(
        axes[0][0], foreign_spot, "net_amount",
        "海外投資家 現物Zスコア推移（過去における偏差の大きさ）"
    )
    if has_futures:
        _draw_zscore_panel(
            axes[1][0], fut_rows, "net_amount_oku",
            "海外投資家 先物Zスコア推移（日経225＋TOPIX合算）"
        )

    fig.tight_layout(pad=1.5)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def build_all_years():
    """全年分のExcelを年ごとに再生成する（スタンドアロン実行用）"""
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / "config" / ".env")
    from db import supabase_client as db
    from scripts.analyze_jpx import INVESTORS

    # 全データ取得
    all_spot, all_futures, all_monthly = [], [], []
    for inv in INVESTORS:
        all_spot    += db.fetch_spot_history(inv, 260)
        all_futures += db.fetch_futures_history(inv, 260)
    all_monthly = db.fetch_monthly_summary(60)
    all_stats   = db.fetch_stats_history("spot", 260)

    # 年ごとに分けて出力
    years = sorted({r["week_date"][:4] for r in all_spot})
    out_dir = Path(__file__).resolve().parent.parent / "outputs" / "excel"
    out_dir.mkdir(parents=True, exist_ok=True)

    for year in years:
        spot_y    = [r for r in all_spot    if r["week_date"][:4] == year]
        futures_y = [r for r in all_futures if r["week_date"][:4] == year]
        monthly_y = [r for r in all_monthly if r["year_month"][:4] == year]
        stats_y   = [r for r in all_stats   if r["week_date"][:4] == year]
        path = out_dir / f"jpx_investor_{year}.xlsx"
        build_excel(spot_y, futures_y, monthly_y, path,
                    all_spot_history=all_spot, all_futures_history=all_futures)
        print(f"  {path.name}: 現物{len({r['week_date'] for r in spot_y})}週, "
              f"月次{len({r['year_month'] for r in monthly_y})}ヶ月")


def build_excel(spot_history: list[dict], futures_history: list[dict],
                monthly_data: list[dict], output_path: Path,
                all_spot_history: list[dict] = None,
                all_futures_history: list[dict] = None) -> Path:
    """Excelファイルを生成して保存
    spot_history: 当該年のみのデータ（行表示用）
    all_spot_history: 全年データ（Zスコア計算用・52週窓を正確に確保するため）
    all_futures_history: 全年先物データ（先物Zスコアチャート用）
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Zスコア計算は全履歴を使う（年フィルター外のデータも必要）
    zscore_base = all_spot_history if all_spot_history else spot_history

    # 現物週次（Zスコア列付き）
    ws_spot = wb.create_sheet("現物_週次")
    _write_spot_sheet(ws_spot, spot_history, futures_history, zscore_base)

    # チャート（外国人ネット・全投資家比較）
    _write_chart_sheet(wb, spot_history)

    # Zスコア推移チャート（全履歴でローリング計算）
    if spot_history:
        ws_z = wb.create_sheet("チャート_Zスコア")
        ws_z.sheet_view.showGridLines = False
        buf_z = _make_zscore_image(zscore_base, all_futures_history)
        if buf_z:
            ws_z.add_image(XLImage(buf_z), "A1")

    # 月次サマリー
    if monthly_data:
        ws_monthly = wb.create_sheet("月次サマリー")
        _write_monthly_sheet(ws_monthly, monthly_data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"[Excel] 保存完了: {output_path}")
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    print("全年分のExcelを再生成します...")
    build_all_years()
    print("完了")
